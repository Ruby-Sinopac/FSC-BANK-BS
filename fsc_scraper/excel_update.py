"""直接把新月份 append 進使用者既有的 Excel（保留其格式、分頁、表頭）。

設計重點：
- 不重寫整張表，只在每個分頁底部新增缺少的月份列。
- 自動辨識：表頭列(含「統計期」)、統計期欄、各項目欄、資料起始列。
- 銀行分頁名以「子字串」對應到網站銀行名（玉山→808玉山商業銀行…）。
- 期別字串、缺值符號、儲存格樣式皆沿用該檔既有風格。
- 寫入前先備份；支援 dry-run 只預覽不寫入。
"""

from __future__ import annotations

import os
import re
import shutil
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime

import openpyxl

from .export import to_number
from .periods import format_roc, parse_period

_PERIOD_CELL = re.compile(r"\d{2,3}\s*年")
_DASH_CHARS = ("—", "–", "－", "-")


def _norm(s) -> str:
    """正規化字串以利比對：去空白、全形括號轉半形。"""
    s = "" if s is None else str(s)
    s = s.replace("（", "(").replace("）", ")").replace("，", ",")
    return re.sub(r"\s+", "", s).strip()


def _strip_code(bank: str) -> str:
    return re.sub(r"^\d{3}\s*", "", str(bank)).strip()


@dataclass
class SheetLayout:
    header_row: int
    period_col: int
    last_data_row: int
    item_cols: dict[str, int]  # 正規化項目名 -> 欄索引
    raw_headers: dict[int, str]  # 欄索引 -> 原始表頭文字
    period_space: str = " "
    period_zero: bool = False
    dash_char: str = "—"


def detect_layout(ws) -> SheetLayout | None:
    """辨識一個分頁的版面。找不到「統計期」表頭則回 None。"""
    header_row = period_col = None
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(1, min(ws.max_column, 5) + 1):
            if _norm(ws.cell(row=r, column=c).value) == "統計期":
                header_row, period_col = r, c
                break
        if header_row:
            break
    if not header_row:
        return None

    item_cols: dict[str, int] = {}
    raw_headers: dict[int, str] = {}
    for c in range(period_col + 1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is None or str(val).strip() == "":
            continue
        raw_headers[c] = str(val).strip()
        item_cols[_norm(val)] = c

    # 找資料列範圍與期別樣式
    last_data_row = header_row
    period_samples: list[str] = []
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=period_col).value
        if a is not None and _PERIOD_CELL.search(str(a)):
            last_data_row = r
            period_samples.append(str(a))

    space, zero = _detect_period_style(period_samples)
    dash = _detect_dash(ws, header_row + 1, last_data_row, period_col, item_cols)
    return SheetLayout(
        header_row=header_row,
        period_col=period_col,
        last_data_row=last_data_row,
        item_cols=item_cols,
        raw_headers=raw_headers,
        period_space=space,
        period_zero=zero,
        dash_char=dash,
    )


def _detect_period_style(samples: list[str]) -> tuple[str, bool]:
    for s in samples:
        m = re.match(r"\s*\d{2,3}\s*年(\s*)(\d{1,2})\s*月", s)
        if m:
            return m.group(1), len(m.group(2)) == 2
    return " ", False


def _detect_dash(ws, r0: int, r1: int, period_col: int, item_cols: dict[str, int]) -> str:
    cols = list(item_cols.values())
    for r in range(r0, r1 + 1):
        for c in cols:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() in _DASH_CHARS:
                return v.strip()
    return "—"


def _fmt_period(code: int, space: str, zero: bool) -> str:
    y, mn = divmod(code, 100)
    mm = f"{mn:02d}" if zero else f"{mn}"
    return f"{y}年{space}{mm}月"


def _safe_parse(v) -> int | None:
    try:
        return parse_period(v)
    except (ValueError, TypeError):
        return None


def _existing_periods(ws, lay: SheetLayout) -> set[int]:
    out: set[int] = set()
    for r in range(lay.header_row + 1, lay.last_data_row + 1):
        code = _safe_parse(ws.cell(row=r, column=lay.period_col).value)
        if code is not None:
            out.add(code)
    return out


def map_bank(sheet_name: str, banks: list[str]) -> str | None:
    """以分頁名為子字串對應到網站銀行名。玉山->808玉山商業銀行。"""
    key = _norm(sheet_name)
    for b in banks:
        if key and key in _norm(_strip_code(b)):
            return b
    return None


@dataclass
class SheetReport:
    sheet: str
    bank: str | None = None
    last_period: int | None = None
    new_periods: list[int] = field(default_factory=list)
    matched_items: int = 0
    unmatched_items: list[str] = field(default_factory=list)
    note: str = ""


def latest_period_in_workbook(path: str) -> int | None:
    """掃描整個活頁簿，回傳所有分頁中最大的期碼。"""
    if not os.path.exists(path):
        return None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    best: int | None = None
    for ws in wb.worksheets:
        lay = detect_layout(ws)
        if not lay:
            continue
        for code in _existing_periods(ws, lay):
            if best is None or code > best:
                best = code
    wb.close()
    return best


def append_to_workbook(long_df, path: str, *, dry_run: bool = False, backup: bool = True, backup_dir: str = "") -> list[SheetReport]:
    """把長格式資料中「各分頁尚缺的月份」append 到既有 Excel。

    backup_dir 留空時，備份存到「目前工作目錄」(即執行程式/批次檔的資料夾)。
    """
    wb = openpyxl.load_workbook(path)
    banks = list(dict.fromkeys(long_df["銀行"].tolist()))
    reports: list[SheetReport] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rep = SheetReport(sheet=sheet)
        lay = detect_layout(ws)
        if not lay:
            rep.note = "找不到『統計期』表頭，略過"
            reports.append(rep)
            continue
        bank = map_bank(sheet, banks)
        rep.bank = bank
        if not bank:
            rep.note = "對不到網站銀行，略過"
            reports.append(rep)
            continue

        sub = long_df[long_df["銀行"] == bank]
        existing = _existing_periods(ws, lay)
        rep.last_period = max(existing) if existing else None

        # 此分頁要新增的期別 = 抓到的、且尚未存在的
        sub_codes = sorted({int(c) for c in sub["期碼"].tolist()})
        new_codes = [c for c in sub_codes if c not in existing]
        rep.new_periods = new_codes

        # 比對項目欄
        scraped_items = list(dict.fromkeys(sub["項目"].tolist()))
        item_to_col: dict[str, int] = {}
        for it in scraped_items:
            col = lay.item_cols.get(_norm(it))
            if col:
                item_to_col[it] = col
            else:
                rep.unmatched_items.append(it)
        rep.matched_items = len(item_to_col)

        if not dry_run and new_codes:
            _write_new_rows(ws, lay, sub, new_codes, item_to_col)
        reports.append(rep)

    if not dry_run and any(r.new_periods and r.bank for r in reports):
        if backup:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            base = os.path.splitext(os.path.basename(path))[0]
            folder = backup_dir or os.getcwd()  # 留空 = 執行程式的資料夾
            os.makedirs(folder, exist_ok=True)
            bak = os.path.join(folder, f"{base}.backup_{ts}.xlsx")
            shutil.copy2(path, bak)
            print(f"已備份原檔 -> {bak}")
        wb.save(path)
    wb.close()
    return reports


def _write_new_rows(ws, lay: SheetLayout, sub, new_codes: list[int], item_to_col: dict[str, int]) -> None:
    row = lay.last_data_row
    # 以最後一筆資料列作為樣式來源
    style_src_row = lay.last_data_row
    for code in new_codes:
        row += 1
        period_str = _fmt_period(code, lay.period_space, lay.period_zero)
        _set_cell(ws, row, lay.period_col, period_str, style_src_row)
        rows = sub[sub["期碼"] == code]
        # 該期各項目值
        vals = {r["項目"]: r["數值"] for _, r in rows.iterrows()}
        for item, col in item_to_col.items():
            v = to_number(vals.get(item))
            if v is None or (isinstance(v, str) and v.strip() in _DASH_CHARS) or v == "":
                v = lay.dash_char
            _set_cell(ws, row, col, v, style_src_row, numeric_col=col)
    lay.last_data_row = row


def _set_cell(ws, row: int, col: int, value, style_src_row: int, numeric_col: int | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    src = ws.cell(row=style_src_row, column=col)
    # 沿用上一列同欄的樣式
    if src.has_style:
        cell.font = copy(src.font)
        cell.border = copy(src.border)
        cell.fill = copy(src.fill)
        cell.alignment = copy(src.alignment)
        cell.number_format = src.number_format
    # 數字欄若樣式來源是文字(—)導致無千分位，補上 #,##0
    if isinstance(value, (int, float)) and cell.number_format in ("General", "@"):
        cell.number_format = "#,##0"
