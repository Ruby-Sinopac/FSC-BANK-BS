"""命令列入口： python -m fsc_scraper <command>

指令：
  discover-menu   抓網站導覽，列出所有節點與 funid，找出「資產負債簡表(102年以後)」
  inspect         dump 某 funid 的查詢條件頁（表單欄位、可選期間、範例網址）
  update          依 config.yaml 抓最新資料並增量合併到歷史檔
"""

from __future__ import annotations

import argparse
import logging
import os
import sys

from . import storage
from .client import StatisClient
from .config import load_config
from .discover import (
    discover_menu,
    inspect_condition_page,
    print_condition_page,
    print_menu,
    save_debug_html,
)
from .periods import current_roc_period, format_roc, next_month, parse_period
from .scraper import build_result_url, detect_latest_period, fetch


def _make_client(cfg) -> StatisClient:
    http = cfg.http
    return StatisClient(
        user_agent=http.get("user_agent", "Mozilla/5.0"),
        timeout=int(http.get("timeout", 60)),
        retries=int(http.get("retries", 4)),
        delay=float(http.get("delay_between_requests", 1.5)),
        encoding=cfg.encoding,
        verify_ssl=cfg.verify_ssl,
    )


def cmd_discover_menu(cfg, args) -> int:
    client = _make_client(cfg)
    print(f"正在抓取網站導覽 ({cfg.base_url}) …")
    nodes = discover_menu(client, cfg.base_url, keyword=args.keyword, debug=args.debug)
    print_menu(nodes, keyword=args.keyword)
    return 0


def cmd_inspect(cfg, args) -> int:
    funid = args.funid or cfg.funid
    if not funid:
        print("請用 --funid 指定，或先在 config.yaml 填 query.funid", file=sys.stderr)
        return 2
    client = _make_client(cfg)
    print(f"正在抓取 funid={funid} 的查詢條件頁 …")
    html, fields = inspect_condition_page(client, cfg.base_url, funid)
    if args.debug:
        path = save_debug_html(html, f"condition_{funid}.html")
        print(f"(原始 HTML 已存至 {path})")
    period_opts = print_condition_page(fields)
    if period_opts:
        latest = max(parse_period(v) for v, _ in period_opts if v.isdigit())
        earliest = min(parse_period(v) for v, _ in period_opts if v.isdigit())
        print(f"\n期間範圍：{format_roc(earliest)} ~ {format_roc(latest)}（最新期 = {latest}）")
        # 用設定的樣板組一條範例網址，方便核對
        tmp_cfg = cfg
        url = build_result_url(tmp_cfg, latest, latest) if funid == cfg.funid else "(funid 與 config 不同，略過範例網址)"
        print(f"\n範例結果網址（最新一期）：\n  {url}")
        print("\n請在瀏覽器開啟上面網址核對是否為正確的『資產負債簡表(102年以後)』；")
        print("若參數不符，請對照實際查詢後的網址，修改 config.yaml 的 result_url_template。")
    return 0


def cmd_update(cfg, args) -> int:
    urls = cfg.download_urls
    if not urls and not cfg.funid:
        print(
            "config.yaml 尚未設定 query.download_urls（建議）或 query.funid。\n"
            "請在瀏覽器查到該表後，把結果頁網址貼到 download_urls。",
            file=sys.stderr,
        )
        return 2
    if urls:
        print(f"共 {len(urls)} 條下載網址。")
    client = _make_client(cfg)

    # 1) 結束期（上界）— CLI --end 優先
    if args.end:
        end = parse_period(args.end)
    elif cfg.end_period == "latest":
        if urls:
            # 用真實下載網址時，以系統當月為上界，實際最新期由回傳資料決定。
            end = current_roc_period()
            print(f"以當月 {format_roc(end)} 為查詢上界（實際最新期以網站回傳為準）。")
        else:
            end = detect_latest_period(client, cfg)
    else:
        end = parse_period(cfg.end_period)

    # 2) 起始期 — CLI --start 優先；否則以「目標 Excel / 主檔」最後一期推算
    existing = storage.load_existing(cfg.data_file, cfg.sheet_name)
    if args.start:
        start = parse_period(args.start)
    elif cfg.start_period == "auto":
        last = None
        if cfg.excel_target_enabled and os.path.exists(cfg.excel_target_file):
            from .excel_update import latest_period_in_workbook
            last = latest_period_in_workbook(cfg.excel_target_file)
            if last is not None:
                print(f"你的 Excel 最後一期：{format_roc(last)}")
        if last is None:
            last = storage.latest_period_in(existing, cfg.period_column)
        if last is None:
            start = 10201
            print("找不到既有資料，將從 102年01月 開始全量抓取。")
        else:
            start = next_month(last)
            print(f"將從 {format_roc(start)} 開始增量抓取至 {format_roc(end)}。")
    else:
        start = parse_period(cfg.start_period)

    if start > end:
        print(f"已是最新（最後一期之後沒有新資料）。最新期 = {format_roc(end)}，無需更新。")
        return 0

    # 3) 抓取（長格式：統計期 / 期碼 / 銀行 / 項目 / 數值）
    result = fetch(client, cfg, start, end, debug=args.debug)
    df = result.df
    print(f"\n解析出 {df.shape[0]} 筆長格式資料。")
    if not df.empty:
        print(f"  期別 ({df['統計期'].nunique()})：", sorted(df['統計期'].unique()))
        print(f"  銀行 ({df['銀行'].nunique()})：", list(dict.fromkeys(df['銀行']))[:20])
        print(f"  項目 ({df['項目'].nunique()})：", list(dict.fromkeys(df['項目']))[:30])
    else:
        print("⚠ 沒有解析到任何資料，請用 --debug 後以 analyze 檢視 HTML。")
        return 1

    # 4) 更新長格式主檔（工具自己的備份來源）
    merged = storage.merge(existing, df, cfg.period_column, cfg.key_columns)
    if not args.dry_run:
        storage.save(merged, cfg.data_file, cfg.sheet_name)
        print(f"已更新主檔 {cfg.data_file}（共 {merged.shape[0]} 筆）。")

    # 5) 直接更新進你原本的 Excel（option 2）
    if cfg.excel_target_enabled:
        from .excel_update import append_to_workbook
        if not os.path.exists(cfg.excel_target_file):
            print(f"⚠ 找不到目標 Excel：{cfg.excel_target_file}")
            return 2
        mode = "[dry-run 預覽] " if args.dry_run else ""
        print(f"\n{mode}更新你的 Excel：{cfg.excel_target_file}")
        reports = append_to_workbook(
            merged, cfg.excel_target_file, dry_run=args.dry_run,
            backup=cfg.excel_target_backup, backup_dir=cfg.excel_target_backup_dir,
        )
        _print_excel_reports(reports)

    # 6) （可選）另外匯出全新的各銀行分頁 Excel
    if cfg.export_enabled and not args.dry_run:
        from . import export
        n = export.export_per_bank(
            merged, cfg.export_file,
            short_names=cfg.short_sheet_names, include_total=cfg.include_total,
        )
        print(f"已另存各銀行分頁：{cfg.export_file}（{n} 個分頁）。")
    return 0


def _print_excel_reports(reports) -> None:
    print(f"{'分頁':<8}{'對應銀行':<18}{'最後一期':<10}{'新增期數':<8}{'項目對到':<8} 備註")
    for r in reports:
        last = format_roc(r.last_period) if r.last_period else "-"
        new = f"{len(r.new_periods)}" if r.bank else "-"
        items = f"{r.matched_items}" if r.bank else "-"
        bank = (r.bank or "")[:16]
        print(f"{r.sheet:<8}{bank:<18}{last:<10}{new:<8}{items:<8} {r.note}")
        if r.new_periods:
            from .periods import format_roc as _f
            print(f"        └ 新增：{[_f(c) for c in r.new_periods]}")
        if r.unmatched_items:
            print(f"        └ ⚠ 對不到的項目欄：{r.unmatched_items}")


def cmd_analyze(cfg, args) -> int:
    """分析一個已存的結果 HTML 檔，列出所有表格的形狀與預覽，協助定位資料表。"""
    import io as _io

    import pandas as pd

    from .scraper import _leaf_tables_html, _table_score, _clean_table

    with open(args.htmlfile, "r", encoding="utf-8", errors="replace") as f:
        text = f.read()
    print(f"檔案：{args.htmlfile}（{len(text)} 字元）\n")

    leaves = _leaf_tables_html(text)
    print(f"最內層(leaf)表格數：{len(leaves)}")
    all_tables: list[tuple[str, "pd.DataFrame"]] = []
    for i, html in enumerate(leaves):
        try:
            for df in pd.read_html(_io.StringIO(html)):
                all_tables.append((f"leaf#{i}", df))
        except ValueError:
            continue

    if not all_tables:
        print("(leaf 解析不到表格，改列整頁所有表格)")
        try:
            for j, df in enumerate(pd.read_html(_io.StringIO(text))):
                all_tables.append((f"all#{j}", df))
        except ValueError:
            pass

    ranked = sorted(all_tables, key=lambda t: _table_score(t[1]), reverse=True)
    print(f"共解析出 {len(all_tables)} 個表格，依『像資料表』分數排序，前 5 個：\n")
    for tag, df in ranked[:5]:
        print(f"── {tag}  shape={df.shape}  score={_table_score(df):.0f}")
        cleaned = _clean_table(df.copy())
        print("   欄位:", list(cleaned.columns)[:20])
        print(cleaned.head(4).to_string()[:1500])
        print()
    print("請把分數最高、看起來像『統計期 + 各銀行各項目數字』的那個表格內容貼給我。")
    return 0


def cmd_inspect_excel(cfg, args) -> int:
    """讀本機 Excel，把結構（分頁、前幾列內容、合併儲存格）印成文字，方便對齊。"""
    import openpyxl

    wb = openpyxl.load_workbook(args.file, data_only=True)
    print(f"檔案：{args.file}")
    print(f"分頁數：{len(wb.sheetnames)}")
    print("分頁名稱：", wb.sheetnames)

    targets = [args.sheet] if args.sheet else wb.sheetnames[:2]
    for sh in targets:
        if sh not in wb.sheetnames:
            print(f"\n(找不到分頁「{sh}」)")
            continue
        ws = wb[sh]
        print(f"\n===== 分頁「{sh}」 維度={ws.dimensions} 前 {args.rows} 列 / 前 {args.cols} 欄 =====")
        merged = [str(r) for r in ws.merged_cells.ranges][:15]
        if merged:
            print("合併儲存格（前15）：", merged)
        for r in range(1, args.rows + 1):
            vals = []
            for c in range(1, args.cols + 1):
                v = ws.cell(row=r, column=c).value
                vals.append("" if v is None else str(v))
            # 去掉尾端空欄，縮短輸出
            while vals and vals[-1] == "":
                vals.pop()
            print(f"R{r}:", " | ".join(vals))
        # 嘗試找「統計期」欄與最後一筆資料
        print("（請把以上內容貼給我；我會據此辨識表頭列、統計期欄、各項目欄與資料起始列。）")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="fsc_scraper", description="資產負債簡表(102年以後) 自動抓取")
    parser.add_argument("-c", "--config", default=None, help="設定檔路徑（預設 config.yaml）")
    parser.add_argument("-v", "--verbose", action="store_true", help="顯示詳細日誌")
    sub = parser.add_subparsers(dest="command", required=True)

    p_menu = sub.add_parser("discover-menu", help="列出選單節點與 funid")
    p_menu.add_argument("--keyword", default="資產負債簡表", help="要尋找的節點關鍵字")
    p_menu.add_argument("--debug", action="store_true", help="把每頁原始 HTML 存到 debug/")

    p_insp = sub.add_parser("inspect", help="dump 查詢條件頁")
    p_insp.add_argument("--funid", default=None)
    p_insp.add_argument("--debug", action="store_true", help="另存原始 HTML 到 debug/")

    p_upd = sub.add_parser("update", help="抓最新資料並增量合併")
    p_upd.add_argument("--dry-run", action="store_true", help="只抓取與預覽，不寫檔")
    p_upd.add_argument("--debug", action="store_true", help="另存原始 HTML 到 debug/")
    p_upd.add_argument("--start", default=None, help="覆寫起始民國年月（測試用，如 11502）")
    p_upd.add_argument("--end", default=None, help="覆寫結束民國年月（測試用，如 11503）")

    p_ana = sub.add_parser("analyze", help="分析已存的結果 HTML，定位資料表")
    p_ana.add_argument("htmlfile", help="debug/ 下的 HTML 檔路徑")

    p_xls = sub.add_parser("inspect-excel", help="檢視本機 Excel 結構（分頁/表頭/前幾列）")
    p_xls.add_argument("file", help="你的歷史 Excel 路徑")
    p_xls.add_argument("--sheet", default=None, help="只看指定分頁（預設看前兩個）")
    p_xls.add_argument("--rows", type=int, default=8, help="顯示前幾列（預設8）")
    p_xls.add_argument("--cols", type=int, default=30, help="顯示前幾欄（預設30）")

    args = parser.parse_args(argv)
    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    try:
        cfg = load_config(args.config)
    except (ValueError, FileNotFoundError) as exc:
        print(f"\n[設定檔錯誤] {exc}", file=sys.stderr)
        return 2

    handlers = {
        "discover-menu": cmd_discover_menu,
        "inspect": cmd_inspect,
        "update": cmd_update,
        "analyze": cmd_analyze,
        "inspect-excel": cmd_inspect_excel,
    }
    return handlers[args.command](cfg, args)


if __name__ == "__main__":
    raise SystemExit(main())
